import os
import re
import collections
import openpyxl
import shutil
import sys
import subprocess
from openpyxl.styles import NamedStyle, Font, Border, Side
from paramiko import SSHClient
from scp import SCPClient

########Declared Variables#######
plat_data ={}
parser_data ={}
release = sys.argv[1]
shipment = sys.argv[2]
feature = sys.argv[3]
zid = sys.argv[4]
usr_pass = sys.argv[5]
rebuild = sys.argv[6]

#####################################
########Functions Definition#########
#####################################

#######Function to check and create lock file######
def check_lock():
        flag=0
        file_list = [lck_file for lck_file in os.listdir(".") if lck_file.endswith('.lck')]
        for item in file_list:
                if shipment in item and rebuild == "true":
                        re_file = [lck_file for lck_file in os.listdir(".") if lck_file.startswith('rebuild')]
                        for re_item in re_file:
                                if shipment in re_item:
                                        flag+=1
                                        print re_item
                                        print "ERROR: PID is already rebuild for "+shipment
                                        exit(1)
                        if flag == 0:
                                print "Rebuilding PID"
                                flag +=1
                                r = open( "rebuild_"+shipment+"_"+zid+".lck" , "w")
                elif shipment in item and rebuild == "false":
                        print item
                        print "ERROR: PID is already created for "+shipment
                        exit(1)
        if flag==0:
                f = open( shipment+"_"+zid+".lck" , "w")


######Function to skip "I","O","P","Q","R","W" character in Revision#########
def check_char(rev_char):
    count = 1
    if "V" in rev_char or "H" in rev_char:
        count = 2
    elif "N" in rev_char:
        count = 5
    return count

######Function to calculate new Revision######
def revUpdate(rev):
        fst_char = rev[0]
        sec_char =""
        if (len(rev) > 1):
                sec_char = rev[1]
        if rebuild == "true":
                new_rev = rev+"1"
        elif (sec_char =="") and (fst_char != "Z"):
                count = check_char(fst_char)
                new_rev = chr(ord(fst_char)+count)
        elif (sec_char == "") and (fst_char == "Z"):
                new_rev = "AA"
        elif (sec_char != "") and (sec_char != "Z"):
                count = check_char(sec_char)
                inc = chr(ord(sec_char)+count)
                new_rev = fst_char+inc
        elif (sec_char != "") and (sec_char == "Z"):
                count = check_char(fst_char)
                inc1 = chr(ord(fst_char)+count)
                inc2 ="A"
                new_rev = inc1 + inc2
        return new_rev

#####Function to update Revision History Sheet#####
def updateRevHistory():
        revFile = myFile.get_sheet_by_name('Revision History')
        last_row = revFile.max_row
        print last_row
        old_rev = revFile.cell(column=1, row=last_row).value
        print "old_rev ="+old_rev
        old_rev = re.match(r"([a-zA-Z]+)", old_rev).groups()[0]   ###removing numeric value
        new_rev = revUpdate(old_rev)
        print "new_rev ="+new_rev

        fontStyle = NamedStyle(name="fontStyle")
        bd = Side(style='thin', color="000000")
        fontStyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if "fontStyle" in myFile.named_styles:
               del myFile._named_styles[ myFile.style_names.index("fontStyle")]

        revFile.cell(row=last_row+1, column=1).style = fontStyle
        if rebuild == "true":
                revFile.cell(row=last_row+1, column=1).value = "P"+new_rev+"-1"
        else:
                revFile.cell(row=last_row+1, column=1).value = "P"+new_rev+"1"
        revFile.cell(row=last_row+1, column=2).style = fontStyle
        revFile.cell(row=last_row+1, column=2).value = "Package details for "+shipment+" Shipment"

        fontStyle.font = Font(bold=True)
        revFile.cell(row=last_row+2, column=1).style = fontStyle
        revFile.cell(row=last_row+2, column=1).value = new_rev
        revFile.cell(row=last_row+2, column=2).style = fontStyle
        revFile.cell(row=last_row+2, column=2).value = "Sharp revised for "+shipment


#####Function to store pkg details in dictionary#######
def pkg_details(pkg_list):
        dict={}
        for pkg in pkg_list:
                pkg_rstate = re.search(r'_R(.+).zip', pkg).groups()[0]
                pkg_rstate = "R"+pkg_rstate
                pkg_name = re.search('(.+)_'+pkg_rstate , pkg).groups()[0]
                pkg_name = pkg_name+".zip"
                dict[pkg_name] = pkg_rstate
        return dict

#####Function to update data in Platform and Parser sheet######
def excel_update(data_dict, sheet_name):
        sheet = myFile.get_sheet_by_name(sheet_name)
        for rowNum in range(2, sheet.max_row + 1):  #skipping first row of file
                pkgName = sheet.cell(row=rowNum, column=1).value
                pkgName = pkgName.strip()
                if pkgName in data_dict:
                        sheet.cell(row=rowNum, column=2).value = data_dict[pkgName]


#######Function to copy file to Desktop#########
def copy_file(PIDFile):
        ssh = SSHClient()
        ssh.load_system_host_keys()
        ssh.connect('seliius26954.seli.gic.ericsson.se',username=zid,password=usr_pass)
        scp = SCPClient(ssh.get_transport())
        scp.put(PIDFile,PIDFile)
        scp.close()


###########################
########MAIN SCRIPT########
###########################

########Platform and Parser paths entered by the user#########
plat_path = "/net/10.45.192.134/JUMP/ENIQ_STATS/ENIQ_STATS/"+shipment+"_Linux/eniq_base_sw/eniq_sw"
parser_path = "/net/10.45.192.134/JUMP/ENIQ_STATS/ENIQ_STATS/"+feature+"/eniq_parsers"

#####checking lock file######
check_lock()

#######checking if parser path exists###########
parser_exist = os.path.exists(parser_path)

#######taking platform packages list ###########
plat_pkgs= [pkg for pkg in os.listdir(plat_path) if pkg.endswith('.zip')]
plat_data = pkg_details(plat_pkgs)
plat_data = collections.OrderedDict(sorted(plat_data.items()))

######taking parser packages list if exist######
if parser_exist:
        parser_pkgs = [pkg for pkg in os.listdir(parser_path) if pkg.endswith('.zip')]
        parser_data = pkg_details(parser_pkgs)
        parser_data = collections.OrderedDict(sorted(parser_data.items()))

#######Taking backup and Opening original ExcelSheet######
PIDFile = 'Package Information Document for '+release+'.xlsx'
shutil.copy(PIDFile, 'Package Information Document for '+release+'_bkp.xlsx')
myFile = openpyxl.load_workbook(PIDFile)

#####update Revision History function call######
updateRevHistory()

#####update platform and parser content function call######
excel_update(plat_data,'Platform')
if parser_exist:
        excel_update(parser_data,'Parser')

######saving updated file#########
myFile.save(PIDFile)
copy_file(PIDFile)
