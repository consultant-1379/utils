import os
import re
import collections
import openpyxl
import shutil
import sys
import subprocess
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from paramiko import SSHClient
from scp import SCPClient
from datetime import date

data = {}
#print("++++",sys.argv)
release = sys.argv[1]
shipment = sys.argv[2]
feature = sys.argv[3]
zid = sys.argv[4]
usr_pass = sys.argv[5]
previousShip = sys.argv[6]
rebuild = sys.argv[7]
final_list = []

#################################################### Function to skip "I","O","P","Q","R","W" character in Revision ###########################################
def check_char(rev_char):
    count = 1
    if "V" in rev_char or "H" in rev_char:
        count = 2
    elif "N" in rev_char:
        count = 5
    return count

############################################################ Function to calculate new Revision ###############################################################
def revUpdate(rev):
        fst_char = rev[0]
        sec_char =""
        if (len(rev) > 1):
                sec_char = rev[1]
        if rebuild == "true":
                new_rev = rev+"1"
        elif (sec_char =="") and (fst_char != "Z"):
                count = check_char(fst_char)
                new_rev = chr(ord(fst_char)+count)
        elif (sec_char == "") and (fst_char == "Z"):
                new_rev = "AA"
        elif (sec_char != "") and (sec_char != "Z"):
                count = check_char(sec_char)
                inc = chr(ord(sec_char)+count)
                new_rev = fst_char+inc
        elif (sec_char != "") and (sec_char == "Z"):
                count = check_char(fst_char)
                inc1 = chr(ord(fst_char)+count)
                inc2 ="A"
                new_rev = inc1 + inc2
        return new_rev

############################################################## Function to update Revision History Sheet ######################################################
def updateRevHistory():
        revFile = myFile.get_sheet_by_name('Revision History')
        last_row = revFile.max_row
        print last_row
        old_rev = revFile.cell(column=1, row=last_row).value
        print "old_rev ="+old_rev
        old_rev = re.match(r"([a-zA-Z]+)", old_rev).groups()[0]   ###removing numeric value
        new_rev = revUpdate(old_rev)
        print "new_rev ="+new_rev

        fontStyle = NamedStyle(name="fontStyle")
        bd = Side(style='thin', color="000000")
        fontStyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if "fontStyle" in myFile.named_styles:
               del myFile._named_styles[ myFile.style_names.index("fontStyle")]

        revFile.cell(row=last_row+1, column=1).style = fontStyle
        revFile.cell(row=last_row+2, column=1).style = fontStyle
        revFile.cell(row=last_row+2, column=1).value = new_rev+"1"
        revFile.cell(row=last_row+2, column=2).style = fontStyle
        revFile.cell(row=last_row+2, column=2).value = date.today().strftime("%d/%m/%Y")
        revFile.cell(row=last_row+2, column=3).style = fontStyle
        revFile.cell(row=last_row+2, column=3).value = zid.upper()
        revFile.cell(row=last_row+2, column=4).style = fontStyle
        revFile.cell(row=last_row+2, column=4).value = "Updated ES "+shipment+" Feature details"

        fontStyle.font = Font(bold=True)
        revFile.cell(row=last_row+1, column=1).style = fontStyle
        revFile.cell(row=last_row+1, column=1).value = "{}".format(old_rev[1:])
        revFile.cell(row=last_row+1, column=2).style = fontStyle
        revFile.cell(row=last_row+1, column=2).value = date.today().strftime("%d/%m/%Y")
        revFile.cell(row=last_row+1, column=3).style = fontStyle
        revFile.cell(row=last_row+1, column=3).value = zid.upper()
        revFile.cell(row=last_row+1, column=4).style = fontStyle
        revFile.cell(row=last_row+1, column=4).value = "Approved version of "+previousShip+ " FID"

######################################################### Function to store cxc details in list #######################################################
def cxc_details(cxc_list):
        for cxc in cxc_list:
                cxc_value = cxc.split(".")[0]
                final_list.append(cxc_value)
        return final_list

######################################################## Function to update data in Techpack sheet from feature path ##################################
def excel_update(data,sheet_name='Tech Pack'):
        sheet = myFile.get_sheet_by_name(sheet_name)
        sheet.cell(row=1,column=4).value = shipment.split("_")[0]+" R-State in SWGW"
        sheet.cell(row=1,column=5).value = shipment.split("_")[0]+" Revision in SWGW"
        for row_pattern in range(2,sheet.max_row+1):
                sheet.cell(row= row_pattern,column=4).fill = PatternFill(patternType='solid',fgColor='FFFFFF')
                sheet.cell(row= row_pattern,column=5).fill = PatternFill(patternType='solid',fgColor='FFFFFF')
                sheet.cell(row= row_pattern,column=6).value= ""
        for cxc in final_list:
                for row in sheet:
                        for cell in row:
                                if cell.value == cxc:
                                        ch = sheet.cell(row= cell.row,column=4).value
                                        if ch == 'R17A' or ch == 'R17B'or ch == 'R17C' or ch == 'R17D' or ch == 'R17E' or ch == 'R17F' or ch == 'R17G':
                                                sheet.cell(row= cell.row,column=4).value = ch[:-1]+chr(ord(ch[3])+1)
                                        else:
                                                sheet.cell(row=cell.row,column=4).value = 'R17A'
                                        sheet.cell(row= cell.row,column=4).fill = PatternFill(patternType='solid',fgColor='7B83EB')
                                        revision = sheet.cell(row= cell.row,column=5).value
                                        if revision == 'GJ1' or revision == 'GJ2' or revision == 'GJ3' or revision == 'GJ4' or revision == 'GJ5' or revision == 'GJ6':
                                                sheet.cell(row= cell.row,column=5).value = revision[:-1]+chr(ord(revision[2])+1)
                                        else:
                                                sheet.cell(row=cell.row,column=5).value = 'GJ1'
                                        sheet.cell(row= cell.row,column=5).fill = PatternFill(patternType='solid',fgColor='7B83EB')
                                        sheet.cell(row= cell.row,column=6).font = Font(bold=True)
                                        sheet.cell(row= cell.row,column=6).value = "Delivered in "+shipment

#################################################### Function to copy file to Desktop ##################################################################
def copy_file(FIDFile):
        ssh = SSHClient()
        ssh.load_system_host_keys()
        ssh.connect('seliius26954.seli.gic.ericsson.se',username=zid,password=usr_pass)
        scp = SCPClient(ssh.get_transport())
        scp.put(FIDFile,FIDFile)
        scp.close()
                                                          ###########################
                                                          ########MAIN SCRIPT########
                                                          ###########################

################################################### feature paths entered by the user ##################################################################
feature_path = "/proj/toathlone/eniq/ENIQ_S/build/ENIQ_S"+release+"/"+release+".DMTEST/LLSV3/"+feature+"/eniq_base_sw/eniq_techpacks/"
report_path = "/proj/toathlone/eniq/ENIQ_S/build/ENIQ_S"+release+"/"+release+".DMTEST/LLSV3/"+feature+"/eniq_base_sw/eniq_reports/reports/"
parser_path = "/proj/toathlone/eniq/ENIQ_S/build/ENIQ_S"+release+"/"+release+".DMTEST/LLSV3/"+feature+"/eniq_base_sw/eniq_sw/"

################################################## checking if feature/report/parser paths exists #############################################################
feature_exist = os.path.exists(feature_path)
reports_exists = os.path.exists(report_path)
parser_exists = os.path.exists(parser_path)

################################################## taking techpack/report/parser packages list ######################################################################
plat_cxc= [cxc for cxc in os.listdir(feature_path) if cxc.endswith('.zip')]
data = cxc_details(plat_cxc)
if reports_exists:
        report_cxc= [cxc for cxc in os.listdir(report_path) if cxc.endswith('.zip')]
        report_data = cxc_details(report_cxc)
if parser_exists:
        parser_cxc = [cxc for cxc in os.listdir(parser_path) if cxc.endswith('.zip')]
        parser_data = cxc_details(parser_cxc)

##################################################### Taking backup and Opening original ExcelSheet #########################################################
FIDFile = 'Feature Information Document for ENIQ STATS '+shipment+'.xlsx'
shutil.copy(FIDFile,'Feature Information Document for ENIQ STATS '+shipment+' bkp.xlsx')
myFile = openpyxl.load_workbook(FIDFile)

##################################################### update Revision History function call ############################################################
updateRevHistory()

#################################################### update content function call #########################################################################
excel_update(data,sheet_name='Tech Pack')
excel_update(data,sheet_name='Special Features')
if reports_exists:
        excel_update(data,sheet_name=' Report')
if parser_exists:
        excel_update(data,sheet_name='Mandatory Packages')

################################################### saving updated file ####################################################################################
myFile.save(FIDFile)
copy_file(FIDFile)

